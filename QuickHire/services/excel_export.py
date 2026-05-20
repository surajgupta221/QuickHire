import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

def export_results_to_excel(job_title: str, results: list) -> bytes:
    """
    Export screening results to a formatted Excel file
    Returns bytes that can be downloaded
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Screening Results"

    # ─── Styles ───────────────────────────────
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill("solid", fgColor="1B4F9E")
    center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ─── Title Row ────────────────────────────
    ws.merge_cells('A1:K1')
    title_cell = ws['A1']
    title_cell.value = f"QuickHire — Screening Results: {job_title}"
    title_cell.font = Font(bold=True, size=14, color="1B4F9E")
    title_cell.alignment = center
    ws.row_dimensions[1].height = 30

    # ─── Headers ──────────────────────────────
    headers = [
        "Rank", "Candidate", "Score", "Match %",
        "Experience", "Education", "Recommendation",
        "Skills Matched", "Skills Missing",
        "Strengths", "Interview Questions"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    ws.row_dimensions[2].height = 20

    # ─── Data Rows ────────────────────────────
    score_colors = {
        range(80, 101): "C6EFCE",   # Green — excellent
        range(60, 80):  "FFEB9C",   # Yellow — good
        range(40, 60):  "FFCC99",   # Orange — fair
        range(0, 40):   "FFC7CE",   # Red — poor
    }

    def get_color(score):
        for r, color in score_colors.items():
            if score in r:
                return color
        return "FFFFFF"

    for row_idx, result in enumerate(results, 3):
        score = result.get("overall_score", 0)
        row_color = get_color(int(score))
        row_fill = PatternFill("solid", fgColor=row_color)

        data = [
            result.get("rank", row_idx - 2),
            result.get("candidate_name", "Unknown"),
            f"{score}/100",
            f"{result.get('match_percentage', 0)}%",
            result.get("experience_match", ""),
            result.get("education_match", ""),
            result.get("recommendation", ""),
            ", ".join(result.get("skills_matched", [])),
            ", ".join(result.get("skills_missing", [])),
            ", ".join(result.get("strengths", [])[:2]),
            result.get("interview_questions", [""])[0] if result.get("interview_questions") else ""
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        ws.row_dimensions[row_idx].height = 40

    # ─── Column Widths ────────────────────────
    column_widths = [8, 25, 10, 10, 15, 15, 20, 35, 35, 30, 40]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ─── Save to bytes ────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()