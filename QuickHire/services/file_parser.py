import PyPDF2
import docx
import openpyxl
import io
from fastapi import UploadFile

async def extract_text_from_file(file: UploadFile) -> str:
    """
    Extract text from PDF, Word, or Excel file
    Returns plain text string
    """
    content = await file.read()
    filename = file.filename.lower()

    # ─── PDF File ─────────────────────────────
    if filename.endswith('.pdf'):
        return extract_from_pdf(content)

    # ─── Word File ────────────────────────────
    elif filename.endswith('.docx') or filename.endswith('.doc'):
        return extract_from_word(content)

    # ─── Excel File ───────────────────────────
    elif filename.endswith('.xlsx') or filename.endswith('.xls'):
        return extract_from_excel(content)

    # ─── Plain Text ───────────────────────────
    elif filename.endswith('.txt'):
        return content.decode('utf-8', errors='ignore')

    else:
        raise ValueError(f"Unsupported file type: {filename}")


def extract_from_pdf(content: bytes) -> str:
    """Extract text from PDF bytes"""
    text = ""
    try:
        pdf_reader = PyPDF2.PdfReader(io.BytesIO(content))
        for page in pdf_reader.pages:
            text += page.extract_text() + "\n"
    except Exception as e:
        text = f"Error reading PDF: {str(e)}"
    return text.strip()


def extract_from_word(content: bytes) -> str:
    """Extract text from Word document bytes"""
    text = ""
    try:
        doc = docx.Document(io.BytesIO(content))
        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"
    except Exception as e:
        text = f"Error reading Word file: {str(e)}"
    return text.strip()


def extract_from_excel(content: bytes) -> str:
    """Extract text from Excel file bytes"""
    text = ""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            text += f"Sheet: {sheet}\n"
            for row in ws.iter_rows(values_only=True):
                row_text = " | ".join(
                    str(cell) for cell in row if cell is not None
                )
                if row_text:
                    text += row_text + "\n"
    except Exception as e:
        text = f"Error reading Excel file: {str(e)}"
    return text.strip()